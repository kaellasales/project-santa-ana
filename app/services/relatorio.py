import openpyxl
from io import BytesIO
from datetime import datetime
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment
from app.repositories.relatorio import RelatorioRepository


class RelatorioService:
    def __init__(self, repository: RelatorioRepository):
        self.repository = repository

    def relatorio_vendas(self, db: Session, data_inicio: datetime, data_fim: datetime):
        resumo = self.repository.resumo_vendas(db, data_inicio, data_fim)
        por_forma = self.repository.por_forma_pagamento(db, data_inicio, data_fim)
        mais_vendidos = self.repository.produtos_mais_vendidos(db, data_inicio, data_fim)

        forma_pagamento = {"DINHEIRO": 0.0, "CARTAO_DEBITO": 0.0, "CARTAO_CREDITO": 0.0, "PIX": 0.0}
        for item in por_forma:
            forma_pagamento[item.tipo.value] = float(item.total or 0)

        return {
            "total_vendas": float(resumo.total_vendas or 0),
            "num_transacoes": resumo.num_transacoes or 0,
            "ticket_medio": float(resumo.ticket_medio or 0),
            "por_forma_pagamento": forma_pagamento,
            "produtos_mais_vendidos": [
                {"nome": p.nome, "quantidade_total": p.quantidade_total}
                for p in mais_vendidos
            ]
        }

    def relatorio_estoque(self, db: Session):
        resumo = self.repository.resumo_estoque(db)
        listagem = self.repository.listagem_produtos(db)

        return {
            "produtos_ativos": resumo["produtos_ativos"],
            "inativos": resumo["inativos"],
            "estoque_baixo": resumo["estoque_baixo"],
            "listagem": [
                {
                    "id": p.id,
                    "nome": p.nome,
                    "categoria": p.categoria_nome,
                    "estoque": p.estoque,
                    "estoque_minimo": p.estoque_minimo,
                    "preco_compra": float(p.preco_compra) if p.preco_compra else None,
                    "preco_venda": float(p.preco_venda)
                }
                for p in listagem
            ]
        }

    def relatorio_margem(self, db: Session, data_inicio: datetime, data_fim: datetime):
        produtos = self.repository.margem_por_produto(db, data_inicio, data_fim)

        receita_bruta = sum(float(p.receita or 0) for p in produtos)
        custo_total = sum(float(p.custo or 0) for p in produtos)
        lucro_bruto = receita_bruta - custo_total
        margem_percentual = round((lucro_bruto / receita_bruta) * 100, 2) if receita_bruta > 0 else 0

        return {
            "receita_bruta": receita_bruta,
            "custo_total": custo_total,
            "lucro_bruto": lucro_bruto,
            "margem_percentual": margem_percentual,
            "detalhamento": [
                {
                    "nome": p.nome,
                    "qtd_vendida": p.qtd_vendida,
                    "receita": float(p.receita or 0),
                    "custo": float(p.custo or 0),
                    "lucro": float(p.receita or 0) - float(p.custo or 0),
                    "margem_percentual": round(((float(p.receita or 0) - float(p.custo or 0)) / float(p.receita or 1)) * 100, 2)
                }
                for p in produtos
            ]
        }

    def relatorio_caixa(self, db: Session, data_inicio: datetime, data_fim: datetime):
        turnos = self.repository.resumo_caixa(db, data_inicio, data_fim)

        total_faturado = sum(t.total_vendas or 0 for t in turnos)
        diferenca_total = sum(t.diferenca or 0 for t in turnos)

        return {
            "turnos_no_periodo": len(turnos),
            "total_faturado": total_faturado,
            "diferenca_total": diferenca_total,
            "historico": [
                {
                    "id": t.id,
                    "abertura": t.data_abertura,
                    "fechamento": t.data_fechamento,
                    "valor_inicial": t.valor_inicial,
                    "total_faturado": t.total_vendas,
                    "diferenca": t.diferenca,
                    "status": t.status.value
                }
                for t in turnos
            ]
        }

    def relatorio_geral(self, db: Session, data_inicio: datetime, data_fim: datetime):
        resumo = self.repository.resumo_vendas(db, data_inicio, data_fim)
        por_forma = self.repository.por_forma_pagamento(db, data_inicio, data_fim)
        mais_vendidos = self.repository.produtos_mais_vendidos(db, data_inicio, data_fim)
        margem = self.repository.margem_por_produto(db, data_inicio, data_fim)

        receita_bruta = sum(float(p.receita or 0) for p in margem)
        custo_total = sum(float(p.custo or 0) for p in margem)
        lucro_bruto = receita_bruta - custo_total

        forma_pagamento = {"DINHEIRO": 0.0, "CARTAO_DEBITO": 0.0, "CARTAO_CREDITO": 0.0, "PIX": 0.0}
        for item in por_forma:
            forma_pagamento[item.tipo.value] = float(item.total or 0)

        produto_top = mais_vendidos[0].nome if mais_vendidos else None

        return {
            "total_vendas": float(resumo.total_vendas or 0),
            "ticket_medio": float(resumo.ticket_medio or 0),
            "lucro_bruto": lucro_bruto,
            "produto_top": produto_top,
            "por_forma_pagamento": forma_pagamento,
            "top_produtos": [
                {"nome": p.nome, "quantidade_total": p.quantidade_total}
                for p in mais_vendidos
            ]
        }


    def exportar_vendas_excel(self, db: Session, data_inicio: datetime, data_fim: datetime):
        from openpyxl.styles import PatternFill, Font, Alignment

        resumo = self.repository.resumo_vendas(db, data_inicio, data_fim)
        por_forma = self.repository.por_forma_pagamento(db, data_inicio, data_fim)
        mais_vendidos = self.repository.produtos_mais_vendidos(db, data_inicio, data_fim)

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        titulo_font = Font(bold=True, size=13)
        periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

        def montar_aba(ws, titulo, headers, widths, rows):
            # Título
            ws.cell(row=1, column=1, value=f"{titulo} — {periodo}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            cell = ws.cell(row=1, column=1)
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal="center")

            # Header
            for col, (title, width) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=2, column=col, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = width

            # Dados
            for row_idx, row_data in enumerate(rows, start=3):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Aba Resumo
        ws1 = wb.active
        ws1.title = "Resumo"
        montar_aba(ws1, "Resumo de Vendas",
            ["Total de Vendas", "Nº Transações", "Ticket Médio"], [20, 15, 15],
            [[float(resumo.total_vendas or 0), resumo.num_transacoes or 0, float(resumo.ticket_medio or 0)]]
        )

        # Aba Forma de Pagamento
        ws2 = wb.create_sheet("Forma de Pagamento")
        montar_aba(ws2, "Vendas por Forma de Pagamento",
            ["Forma", "Total"], [20, 15],
            [[item.tipo.value, float(item.total or 0)] for item in por_forma]
        )

        # Aba Produtos Mais Vendidos
        ws3 = wb.create_sheet("Produtos Mais Vendidos")
        montar_aba(ws3, "Produtos Mais Vendidos",
            ["Produto", "Quantidade"], [40, 15],
            [[p.nome, p.quantidade_total] for p in mais_vendidos]
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=relatorio_vendas.xlsx"}
        )

    def exportar_estoque_excel(self, db: Session):
        resumo = self.repository.resumo_estoque(db)
        listagem = self.repository.listagem_produtos(db)

        wb = openpyxl.Workbook()

        # Aba Resumo
        ws1 = wb.active
        ws1.title = "Resumo"
        
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, title in enumerate(["Produtos Ativos", "Inativos", "Estoque Baixo"], 1):
            cell = ws1.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws1.column_dimensions[cell.column_letter].width = 20

        ws1.append([resumo["produtos_ativos"], resumo["inativos"], resumo["estoque_baixo"]])

        # Aba Listagem
        ws2 = wb.create_sheet("Produtos")
        headers = ["ID", "Nome", "Categoria", "Estoque", "Estoque Mínimo", "Status", "Preço Compra", "Preço Venda"]
        widths = [8, 40, 25, 12, 15, 20, 15, 15]

        for col, (title, width) in enumerate(zip(headers, widths), 1):
            cell = ws2.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[cell.column_letter].width = width

        for p in listagem:
            status = "⚠️ Estoque Baixo" if p.estoque <= p.estoque_minimo else "✅ Normal"
            ws2.append([
                p.id, p.nome, p.categoria_nome, p.estoque, p.estoque_minimo,
                status,
                float(p.preco_compra) if p.preco_compra else None,
                float(p.preco_venda)
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=relatorio_estoque.xlsx"}
        )


    def exportar_margem_excel(self, db: Session, data_inicio: datetime, data_fim: datetime):
        from openpyxl.styles import PatternFill, Font, Alignment

        produtos = self.repository.margem_por_produto(db, data_inicio, data_fim)

        receita_bruta = sum(float(p.receita or 0) for p in produtos)
        custo_total = sum(float(p.custo or 0) for p in produtos)
        lucro_bruto = receita_bruta - custo_total
        margem_percentual = round((lucro_bruto / receita_bruta) * 100, 2) if receita_bruta > 0 else 0

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        titulo_font = Font(bold=True, size=13)
        periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

        def montar_aba(ws, titulo, headers, widths, rows):
            ws.cell(row=1, column=1, value=f"{titulo} — {periodo}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            cell = ws.cell(row=1, column=1)
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal="center")

            for col, (title, width) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=2, column=col, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = width

            for row_idx, row_data in enumerate(rows, start=3):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Aba Resumo
        ws1 = wb.active
        ws1.title = "Resumo"
        montar_aba(ws1, "Análise de Margem",
            ["Receita Bruta", "Custo Total", "Lucro Bruto", "Margem %"], [18, 15, 15, 12],
            [[receita_bruta, custo_total, lucro_bruto, margem_percentual]]
        )

        # Aba Detalhamento
        ws2 = wb.create_sheet("Detalhamento por Produto")
        montar_aba(ws2, "Detalhamento por Produto",
            ["Produto", "Qtd Vendida", "Receita", "Custo", "Lucro", "Margem %"], [40, 12, 15, 15, 15, 12],
            [
                [
                    p.nome,
                    p.qtd_vendida,
                    float(p.receita or 0),
                    float(p.custo or 0),
                    float(p.receita or 0) - float(p.custo or 0),
                    round(((float(p.receita or 0) - float(p.custo or 0)) / float(p.receita or 1)) * 100, 2)
                ]
                for p in produtos
            ]
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=relatorio_margem.xlsx"}
        )